from django.shortcuts import render
from django.views.generic import ListView
from django.urls import reverse
from .forms import AddSemesterForm
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import openpyxl
from .models import (User,
                     Grade,
                     Student,
                     Semester,
                     Professor,
                     Notification,
                     Industry)


# Create your views here.
@login_required
def user_logout(request):
    logout(request)
    return HttpResponseRedirect(reverse('index'))


def manager(request):

    if (request.user and request.user.is_authenticated
            and request.user.is_superuser):

        # declaring template
        template = "manager.html"
        # data = Profile.objects.all()

        # GET request returns the value of the data with the specified key.
        if request.method == "GET":
            form = AddSemesterForm()
            context = {}
            return render(request, template, context)

        if 'upload_user' in request.POST:

            form = AddSemesterForm()
            context = {}

            exl_file = request.FILES.get('exl_file')
            user_type = request.POST['user_type']

            if not exl_file:

                messages.error(request, 'فایل را انتخاب کنید')
                return render(request, template, context)

            elif user_type == '0':
                messages.error(request, 'نوع کاربر را انتخاب کنید')
                return render(request, template, context)

            wb = openpyxl.load_workbook(exl_file)
            worksheet = wb["Sheet1"]

            if user_type == '1':  # For Students

                excel_data = list()
                # iterating over the rows and
                # getting value from each cell in row
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)
                count = 0
                for row in excel_data:

                    if count == 0:
                        count = 1
                        continue

                    if row[0] != 'None':
                        user, created = User.objects.get_or_create(
                            username=row[0],
                            defaults={'first_name': row[1], 'last_name': row[2]},)

                        password = User.objects.make_random_password()

                        user.set_password(password)
                        user.save()

                        student, created = Student.objects.update_or_create(
                            student_id_number=row[3], defaults={'phone_number': row[4],
                                                                'user': user},)

                messages.success(request, 'عملیات انجام شد')
                return render(request, template, context)

            elif user_type == '2':  # Add professor

                excel_data = list()
                # iterating over the rows and
                # getting value from each cell in row
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)
                count = 0
                for row in excel_data:

                    if count == 0:
                        count = 1
                        continue

                    if row[0] != 'None':
                        user, created = User.objects.update_or_create(
                            username=row[0],
                            defaults={'first_name': row[1], 'last_name': row[2]},)

                        password = User.objects.make_random_password()
                        user.set_password(password)
                        user.save()

                        professor, created = Professor.objects.update_or_create(
                            personal_id=row[3],
                            defaults={'phone_number': row[4], 'user': user},)

                messages.success(request, 'عملیات انجام شد')
                return render(request, template, context)

            elif user_type == '3':  # Add industry

                excel_data = list()
                # iterating over the rows and
                # getting value from each cell in row
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)
                count = 0
                for row in excel_data:

                    if count == 0:
                        count = 1
                        continue

                    if row[0] != 'None':

                        user, created = User.objects.update_or_create(
                            username=row[0],
                            defaults={'first_name': row[1], 'last_name': row[2]},)

                        password = User.objects.make_random_password()
                        user.set_password(password)
                        user.save()

                        industry, created = Industry.objects.update_or_create(
                            phone_number=row[3],
                            defaults={'company_name': row[4], 'user': user},)

                messages.success(request, 'عملیات انجام شد')
                return render(request, template, context)

        elif 'add_Semester' in request.POST:
            name = request.POST['semester_name']
            semester, created = Semester.objects.get_or_create(name=name)
            if created:

                messages.success(request, 'عملیات انجام شد')
                form = AddSemesterForm()
                context = {}
                return render(request, template, context)

            else:

                messages.error(request,
                            'نیمسالی با این نام قبلا ایجاد شده است')
                form = AddSemesterForm()
                context = {}
                return render(request, template, context)

        elif 'add_student_to_semester' in request.POST:

            exl_file = request.FILES.get('exl_file2')
            context = {}

            if not exl_file:

                messages.error(request, 'فایل را انتخاب کنید')
                return render(request, template, context)

            wb = openpyxl.load_workbook(exl_file)
            worksheet = wb["Sheet1"]

            excel_data = list()
            # iterating over the rows and
            # getting value from each cell in row
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            count = 0
            for row in excel_data:

                if count == 0:
                    count = 1
                    continue

                student = Student.objects.get(student_id_number=row[0])
                professor = Professor.objects.get(personal_id=row[2])
                semester = Semester.objects.get(name=row[3])

                student.research_subject = row[1]
                student.guid_instructor = professor
                student.semester = semester
                student.save()

            messages.success(request, 'عملیات انجام شد')
            return render(request, template, context)

            return render(request, template, context)

    else:
        context = {}
        template = 'permission_error.html'
        return render(request, template, context)


def index(request):
    # This view will be called for index.html file in
    #   project_management/templates/ directory
    template = 'index.html'

    #  we will use PresentationStudent as model to
    #  display  Presentation Students info on the index.html
    model = Semester

    # The name we use to call presentation_students objects in template
    context_object_name = 'Semester'

    context = {}

    semesters = Semester.objects.all().order_by('-created_date')

    semester = semesters[0]

    Semester_students = semester.students.all()
    context['Semester_students'] = Semester_students



            # else:
            #     grade.value *= 1

    context['notifications'] = Notification.objects.all()
    context['selected'] = semester
    context['semesters'] = semesters

    # if someone tried to login and a post request was sent
    #  the following method will be called
    if request.method == 'POST':

        if 'log_in' in request.POST:

            # Get data from username and password fields of the form
            username = request.POST.get('username')
            password = request.POST.get('password')

            # Check username and password for authentication
            # Password are hashed using argon2 algorithm
            user = authenticate(username=username, password=password)

            if user:
                # The following section will run if
                # the authentication was successful
                if user.is_active:
                    # If the user account is active the user will be logged in
                    login(request, user)
                    # The user will be redirected to the same page(index.html)
                    return HttpResponseRedirect(request.path_info)

            else:
                # print('Failed Login Attempt')
                # print(f"Username: {username} and password {password}")
                # if authentication was Not successful
                #  the following messeage will be displayed
                messages.error(request,
                               '!نام کاربری و یا کلمه عبور وارد شده اشتباه است')
                return HttpResponseRedirect(request.path_info)

        elif 'change' in request.POST:
            semester = Semester.objects.get(name=request.POST['semester_name'])

            context['Semester_students'] = semester.students.all()
            context['selected'] = semester

            return render(request, template, context)

        elif 'submit_grade' in request.POST:
            receiver_student_id = request.POST['to']
            receiver_student = Student.objects.get(id=receiver_student_id)

            if receiver_student.semester.active:
                if request.user.is_authenticated:
                    given_by_user = request.user

                    if Student.objects.filter(user=given_by_user).count() == 1:
                        grade_type = 'by student'

                    elif Industry.objects.filter(user=given_by_user).count() == 1:
                        grade_type = 'by industry'

                    elif Professor.objects.filter(user=given_by_user).count() == 1:

                        prof = Professor.objects.get(user=given_by_user)

                        if prof == receiver_student.guid_instructor:
                            grade_type = 'by guid_instructor'

                        else:
                            grade_type = 'by professor'

                    grade_value = request.POST['grade']

                    if not 0 <= float(grade_value) <= 20:
                        messages.error(request, 'لطفا برای نمره یک عدد بین 0 تا 20 انتخاب کنید')
                        return HttpResponseRedirect(request.path_info)

                    grade, created = Grade.objects.update_or_create(
                        grade_type=grade_type, given_by=given_by_user,
                        given_to=receiver_student.user,
                        defaults={'value': grade_value,
                                  'grade_type': grade_type},)

                    s_grade, p_grade, gi_grade, i_grade = 0, 0, 0, 0
                    s_grade_count, p_grade_count, gi_grade_count, i_grade_count = 0, 0, 0, 0
                    check = False
                    for grade in receiver_student.user.grades.all():

                        # val = grade.value
                        if grade.grade_type == 'by student':
                            check = True

                            s_grade += grade.value
                            s_grade_count += 1

                        elif grade.grade_type == 'by professor':
                            check = True

                            p_grade += grade.value
                            p_grade_count += 1

                        elif grade.grade_type == 'by guid_instructor':
                            check = True

                            gi_grade += grade.value
                            gi_grade_count += 1

                        elif grade.grade_type == 'by industry':
                            check = True

                            i_grade += grade.value
                            i_grade_count += 1

                    if s_grade_count > 0:
                        check = True
                        s_grade /= s_grade_count

                    if p_grade_count > 0:
                        check = True
                        p_grade /= p_grade_count

                    if gi_grade_count > 0:
                        check = True
                        gi_grade /= gi_grade_count

                    if i_grade_count > 0:
                        check = True
                        i_grade /= i_grade_count

                    if check:

                        total = (s_grade*3) + (p_grade*7) + (gi_grade*8) + (i_grade*2)

                        receiver_student.total_grade = total/20
                        receiver_student.save()

                    messages.success(request, 'نمره شما با موفقیت ثبت شد')
                    return HttpResponseRedirect(request.path_info)

            else:
                messages.error(request,
                               'نمی توانید به دانشجویان نیمسالی  که فعال نیست نمره دهید')
                return HttpResponseRedirect(request.path_info)

    return render(request, template, context)
